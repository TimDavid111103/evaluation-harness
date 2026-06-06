from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from eval_harness.models import EvalRow, RunRef
from eval_harness.targets.excel import write_excel


def test_write_excel_creates_grade_ready_sheet(tmp_path: Path):
    rows = [
        EvalRow(
            input="[user]\nHello",
            model_response='{\n  "winner": "no_rag"\n}',
            model_critique="Correct",
            model_outcome="pass",
            trace_id="trace-1",
        )
    ]
    run = RunRef(
        dataset_name="my-dataset",
        run_name="run-1",
        run_id="run-id",
        created_at=datetime.now(timezone.utc),
    )
    path = write_excel(rows, run, tmp_path)

    wb = load_workbook(path)
    ws = wb.active
    assert ws["A1"].value == "Match %"
    assert "AVERAGE" in str(ws["B1"].value)
    assert ws["A2"].value == "Input"
    assert ws["D3"].value == "pass"
    assert not ws["E3"].value
    assert "IF" in str(ws["G3"].value)
