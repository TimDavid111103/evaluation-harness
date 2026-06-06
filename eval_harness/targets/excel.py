from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from eval_harness.models import EvalRow, RunRef

HEADERS = [
    "Input",
    "Model response",
    "Model critique",
    "Model outcome",
    "Human Critique",
    "Human Outcome",
    "Agreement",
]


def _safe_filename(value: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in value)


def write_excel(rows: list[EvalRow], run: RunRef, output_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    filename = (
        f"eval-harness-{_safe_filename(run.dataset_name)}-"
        f"{_safe_filename(run.run_name)}-{timestamp}.xlsx"
    )
    path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Eval harness"

    ws["A1"] = "Match %"
    ws["B1"] = '=IFERROR(AVERAGE(G3:G),"")'

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row_idx, row in enumerate(rows, start=3):
        ws.cell(row=row_idx, column=1, value=row.input)
        ws.cell(row=row_idx, column=2, value=row.model_response)
        ws.cell(row=row_idx, column=3, value=row.model_critique)
        ws.cell(row=row_idx, column=4, value=row.model_outcome)
        ws.cell(row=row_idx, column=5, value="")
        ws.cell(row=row_idx, column=6, value="")
        ws.cell(row=row_idx, column=7, value=f'=IF(F{row_idx}="","",D{row_idx}=F{row_idx})')

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

    last_row = max(3, len(rows) + 2)
    dv = DataValidation(type="list", formula1='"pass,fail"', allow_blank=True)
    dv.add(f"F3:F{last_row}")
    ws.add_data_validation(dv)

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 80 if col <= 3 else 18

    ws.freeze_panes = "A3"
    wb.save(path)
    return path
